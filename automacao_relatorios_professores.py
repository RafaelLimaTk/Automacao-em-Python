from openpyxl.utils import get_column_letter
from tkinter import filedialog, PhotoImage, ttk
from zipfile import ZipFile
from PIL import Image, ImageTk

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import PatternFill, Font, Color, Alignment
from textwrap import wrap

import tkinter.messagebox as messagebox
import pandas as pd
import tkinter as tk
import os
import shutil
import re
import win32com.client
import gc

zip_file_paths = []

class ProfessorEvaluation:
    def __init__(self, professor_name, question, evaluation_data, evaluation_data_value, weighted_average):
        self.professor_name = professor_name
        self.question = question
        self.evaluation_data = evaluation_data
        self.evaluation_data_value = evaluation_data_value
        self.weighted_average = weighted_average

def log_error(message):
    with open("error_log.txt", "a") as f:
        f.write(message + "\n")

def validate_dataframe(df, file_path):
    if df.empty:
        log_error(f"O arquivo {file_path} está vazio.")
        return False
    
    found_categories = set(df.iloc[2].dropna())

    EXPECTED_CATEGORIES = set(["Discordo totalmente", "Discordo", "Nem concordo nem discordo", 
                           "Concordo", "Concordo totalmente", "Não se aplica / não sei responder", 
                           "Total", "Weighted Average"])
    
    if EXPECTED_CATEGORIES != found_categories:
        log_error(f"Categorias incompletas no arquivo {file_path}. Esperado: {EXPECTED_CATEGORIES}, Encontrado: {found_categories}")
        return False
    
    return True

def read_excel_file(input_file_path):
    try:
        df = pd.read_excel(input_file_path, header=None)
        if not validate_dataframe(df, input_file_path):
            return None
        return df
    except FileNotFoundError:
        log_error(f"O arquivo {input_file_path} não foi encontrado.")
        return None
    except PermissionError:
        log_error(f"Permissão negada ao tentar acessar o arquivo {input_file_path}.")
        return None
    except Exception as e:
        log_error(f"Erro ao ler o arquivo {input_file_path}: {e}")
        return None

def create_excel_report_for_professor(professor_evaluation_list, output_file_path, folder_path, filename):

    year, period, course = extract_year_and_period(filename)
    
    append_str = ""
    if course:
        append_str += f"_{course}"
    if year:
        append_str += f"_{year}"
    if period:
        append_str += f"_{period}"

    wb = Workbook()
    ws = wb.active

    for idx, prof_eval in enumerate(professor_evaluation_list):
        if idx > 0:
            ws = wb.create_sheet(title=f"Q{idx+1}")

        output_file_path = f"Avaliação_{format_filename(prof_eval.professor_name)}{append_str}.xlsx"
        
        
        formatted_question = format_question(prof_eval.question)
        ws.append([f"Avaliação Docente {year} {course} | {period} | {prof_eval.professor_name}"])
        ws.append([])
        ws.append(['Média Geral do Professor/Preceptor'])
        ws.append([])
        ws.append([f"Questão: {formatted_question}"])
        ws.append(['Características de Avaliação', 'Porcentagem'])
        ws.append(['Média ponderada', f'Média Ponderada ({prof_eval.weighted_average})'])

        ordered_eval_data_keys = list(prof_eval.evaluation_data.keys())
        ordered_eval_data_values = [prof_eval.evaluation_data_value[i] for i in sorted(prof_eval.evaluation_data_value.keys())]

        for row_idx, (eval_char, eval_value) in enumerate(zip(ordered_eval_data_keys, ordered_eval_data_values), start=8):
            ws.append([eval_char, eval_value])
            if eval_char != 'Total':
                ws.cell(row=row_idx, column=2).number_format = '0.00%'

        chart = BarChart()
        chart.title = prof_eval.question
        chart.x_axis.title = prof_eval.professor_name
        chart.y_axis.title = "Porcentagem"

        chart.width = 23
        chart.height = 10

        data = Reference(ws, min_col=2, min_row=7, max_row=6 + len(prof_eval.evaluation_data), max_col=2)
        cats = Reference(ws, min_col=1, min_row=8, max_row=8 + len(prof_eval.evaluation_data))

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        #config de layout
        fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        noyan_font = Font(name='Noyan Slim', size=12)

        azul_instituicao = Color(rgb="2D4A62")
        fonte_azul = Font(name='Noyan Slim', size=13, color=azul_instituicao)

        for row in ws['A1:C14']:
            for cell in row:
                cell.font = noyan_font

        ws['A6'].fill = fill
        ws['B6'].fill = fill
        ws['A3'].fill = fill
        ws['B3'].fill = fill
        ws['A6'].font = fonte_azul
        ws['B6'].font = fonte_azul
        ws['A3'].font = fonte_azul

        # adjust_column_width_based_on_cell(ws, 'A', 11)
        adjust_column_width_based_on_cell(ws, 'B', 7)
        # adjust_column_width_based_on_cell(ws, 'A', 3)
        ws.column_dimensions['A'].width = 90
        adjust_cell_for_wrapped_text(ws, 'A5')
        #fim da config de layout

        ws['B3'] = get_general_weighted_average(professor_evaluation_list)
        ws.add_chart(chart, f"A{8 + len(prof_eval.evaluation_data)}")

    try:
        save_path = os.path.join(folder_path, output_file_path)
        wb.save(save_path)

        df = pd.read_excel(save_path)

        save_path_absolute = os.path.abspath(save_path)
        pdf_output_file_path = save_path_absolute.replace('.xlsx', '.pdf')
        pdf_output_file_path_absolute = os.path.abspath(pdf_output_file_path)

        conversions = [(save_path_absolute, pdf_output_file_path_absolute)]

        batch_excel_to_pdf(conversions, orientation="Landscape")
    
    except PermissionError:
        log_error(f"Permissão negada ao tentar salvar o arquivo {output_file_path}.")
    except Exception as e:
        log_error(f"Erro ao salvar o arquivo {output_file_path}: {e}")

def batch_excel_to_pdf(conversions, orientation="Portrait"):
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.ScreenUpdating = False
    
    if orientation == "Landscape":
        orientation_const = 2
    else:
        orientation_const = 1

    try:
        for input_file, output_file in conversions:
            if not os.path.exists(input_file):
                print(f"O arquivo {input_file} não existe.")
                continue
            
            workbook = excel.Workbooks.Open(input_file)
            ws_index_list = list(range(1, workbook.Worksheets.Count + 1))
            
            for ws_index in ws_index_list:
                ws = workbook.Worksheets(ws_index)
                ws.PageSetup.Orientation = orientation_const

                ws.PageSetup.LeftMargin = 24 
                ws.PageSetup.RightMargin = 24
                ws.PageSetup.TopMargin = 24
                ws.PageSetup.BottomMargin = 24
            
            workbook.Worksheets(ws_index_list).Select()
            workbook.ActiveSheet.ExportAsFixedFormat(0, output_file)
            workbook.Close(True)

    except Exception as e:
        print(f"Erro na conversão: {e}")
        
    finally:
        excel.Quit()
        del excel
        gc.collect()

def adjust_column_width_based_on_cell(ws, column_letter, row_number):
    cell_value = ws[f"{column_letter}{row_number}"].value
    if cell_value:
        cell_length = len(str(cell_value))
        adjusted_width = cell_length + 2 
        ws.column_dimensions[column_letter].width = adjusted_width

def get_general_weighted_average(prof_evaluations):
    total_weighted_average = 0
    num_questions = len(prof_evaluations)
    
    for eval in prof_evaluations:
        total_weighted_average += eval.weighted_average

    return str(round(total_weighted_average / num_questions, 2))

def format_question(question, max_length=70):
    question = question.replace("Q12.", "").strip()
    wrapped_text = wrap(question, max_length)
    return '\n'.join(wrapped_text)

def adjust_cell_for_wrapped_text(ws, cell_address):
    cell = ws[cell_address]
    cell.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[cell.row].height = 15 * len(cell.value.split('\n'))

def process_excel_file(input_file_path, folder_path):
    df = read_excel_file(input_file_path)
    if df is None:
        log_error(f"Não foi possível processar o arquivo {input_file_path}.")
        return
    
    filename = os.path.basename(input_file_path)
    
    question_rows = df[df.iloc[:, 0].str.contains("Q[0-9]+", na=False, regex=True)].index.tolist()
    all_professors_data = {}
    
    for i in range(len(question_rows) - 1):
        start_row = question_rows[i]
        end_row = question_rows[i + 1] if i < len(question_rows) - 1 else len(df)

        question = df.iloc[start_row, 0]
        evaluation_characteristics = df.iloc[start_row + 1, 1::2].dropna()

        professor_data = df.iloc[start_row + 2 : end_row].dropna(subset=[df.columns[0]])
        professor_data = professor_data.iloc[:, ::1]
        
        eval_chars_list = evaluation_characteristics.tolist()
        odd_indices = list(range(1, len(professor_data.columns), 2))
        for index, row in professor_data.iterrows():
            professor_name = row.iloc[0]
            if not professor_name or professor_name.strip() == "":
                log_error(f"Nome do professor ausente ou inválido no arquivo {input_file_path} na linha {index + 1}.")
                continue
            evaluation_data = {eval_chars_list[i]: row.iloc[i+1] for i in range(len(eval_chars_list))}
            evaluation_data_value = {professor_data.columns[i]: row.iloc[i] for i in odd_indices}

            ordered_values = [evaluation_data_value[i] for i in sorted(evaluation_data_value.keys())]
            for i, key in enumerate(evaluation_data.keys()):
                evaluation_data[key] = ordered_values[i]

            weighted_average = row.iloc[-1]      
            prof_eval = ProfessorEvaluation(professor_name, question, evaluation_data, evaluation_data_value, weighted_average)
            
            if professor_name not in all_professors_data:
                all_professors_data[professor_name] = []
            
            all_professors_data[professor_name].append(prof_eval)

    for professor_name, professor_evaluation_list in all_professors_data.items():
        output_file_path = f"Avaliação_{professor_name.replace('/', '_').replace(' ', '_')}.xlsx"
        create_excel_report_for_professor(professor_evaluation_list, output_file_path, folder_path, filename)

def download_zip_file():
    save_directory = filedialog.askdirectory(title="Escolha o diretório onde os arquivos ZIP serão salvos")
    if not save_directory:
        return
    for zip_file_path in zip_file_paths:
        folder_name = os.path.basename(zip_file_path).replace('.zip', '')
        destination_path = os.path.join(save_directory, f"{folder_name}.zip")
        try:
            shutil.copy(zip_file_path, destination_path)
        except FileNotFoundError:
            log_error(f"O arquivo ZIP {zip_file_path} não foi encontrado.")
        except PermissionError:
            log_error(f"Permissão negada ao tentar copiar o arquivo ZIP {zip_file_path} para {destination_path}.")
        except Exception as e:
            log_error(f"Erro ao copiar o arquivo ZIP {zip_file_path}: {e}")

        os.remove(zip_file_path)
        shutil.rmtree(os.path.join('Avaliações', folder_name))
    
    zip_file_paths.clear()
    download_button.config(state=tk.DISABLED)
    progressbar['value'] = 0
    messagebox.showinfo("Sucesso", "Os arquivos ZIP foram salvos com sucesso.")

def select_and_process_files():
    global root
    global zip_file_paths
    zip_file_paths = []

    file_paths = filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx")])
    if not file_paths:
        return

    total_files = len(file_paths)
    processed_files = 0

    os.makedirs("Avaliações", exist_ok=True)

    for file_path in file_paths:
        if not file_path.lower().endswith('.xlsx'):
            print(f"O arquivo {file_path} não é um arquivo .xlsx válido.")
            continue

        current_folder_name = os.path.splitext(os.path.basename(file_path))[0]
        folder_path = os.path.join("Avaliações", current_folder_name)
        os.makedirs(folder_path, exist_ok=True)

        process_excel_file(file_path, folder_path)

        current_zip_file_path = os.path.join("Avaliações", f"{current_folder_name}.zip")
        with ZipFile(current_zip_file_path, 'w') as zipf:
            for folder_root, _, files in os.walk(folder_path):
                for file in files:
                    zipf.write(os.path.join(folder_root, file), os.path.relpath(os.path.join(folder_root, file), folder_path))

        zip_file_paths.append(current_zip_file_path)

        processed_files += 1
        progress = (processed_files / total_files) * 100
        progressbar['value'] = progress
        root.update_idletasks()
    
    download_button.config(state=tk.NORMAL, command=lambda: download_zip_file())

def extract_year_and_period(filename):
    match = re.search(r'(\d{4}\.\d)\s.*?(\d\s+PERÍODO)', filename)
    year = period = course = None

    if match:
        year, period = match.groups()
    else:
        year_match = re.search(r'(\d{4}\.\d)', filename)
        period_match = re.search(r'(\d\s+PERÍODO)', filename)

        if year_match:
            year = year_match.group(1)
        
        if period_match:
            period = period_match.group(1)
    
    course_match = re.search(r'-\s(.+?)\s-', filename)
    if course_match:
        course = course_match.group(1)

    return year, period, course

def format_filename(filename):
    invalid_chars = ['<', '>', ':', '"', '/', '\\', '|', '?', '*']
    
    for char in invalid_chars:
        filename = filename.replace(char, '_')
    
    filename = '_'.join(filename.split())
    
    return filename

root = tk.Tk()
root.title("Processador de Arquivos Excel")

screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

window_width = 450
window_height = 560

position_x = int((screen_width / 2) - (window_width / 2))
position_y = int((screen_height / 2) - (window_height / 2))

root.geometry(f"{window_width}x{window_height}+{position_x}+{position_y}")
root.configure(bg="#f4fdfe")
root.resizable(False,False)

button_frame = tk.Frame(root)
button_frame.pack(side=tk.TOP, pady=20)
button_frame.configure(bg="#f4fdfe")

img_select_file_button = tk.PhotoImage(file="img/btn_selecionar_arquivo.png")
button = tk.Button(button_frame, image=img_select_file_button, command=select_and_process_files, borderwidth=0, highlightthickness=0, relief='flat')
button.pack(side=tk.LEFT, padx=10)

img_download_file_button = tk.PhotoImage(file="img/btn_baixar_zip.png")
download_button = tk.Button(button_frame, image=img_download_file_button, command=download_zip_file, state=tk.DISABLED, borderwidth=0, highlightthickness=0, relief='flat')
download_button.pack(side=tk.LEFT, padx=10)

progressbar = ttk.Progressbar(root, orient='horizontal', length=420, mode='determinate')
progressbar.pack(pady=20)

blue_frame = tk.Frame(root, bg='#0095F3', height=100)
blue_frame.pack(side=tk.BOTTOM, fill=tk.X)

image_path = "img/FADBA UNIAENE.png"
image = Image.open(image_path)
image = image.resize((160, 60))
photo = ImageTk.PhotoImage(image)

image_label = tk.Label(blue_frame, image=photo, bg='#0095F3')
image_label.pack(side=tk.TOP, pady=5, padx=5) 

root.mainloop()
